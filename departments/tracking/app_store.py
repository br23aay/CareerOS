"""
departments/tracking/app_store.py — the application record + Excel export.

Stores, per application: role, company, contact name, contact LinkedIn,
resume used (filename only), and status. Exports to a real .xlsx you can open
in Excel — exactly the sheet you asked for.
"""

import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent.parent.parent))
from core.base_agent import BaseAgent
from core import config, database


class ApplicationStore(BaseAgent):
    name = "tracking.store"

    def record(self, job_id: int, contact_name: str = "",
               contact_linkedin: str = "", resume_name: str = "",
               status: str = "applied") -> dict:
        """Save/update the rich application record."""
        s = database.get_session()
        app = s.query(database.Application).filter_by(job_id=job_id).first()
        if not app:
            app = database.Application(job_id=job_id)
            s.add(app)
        app.status = status
        if resume_name:
            app.cv_used = resume_name
        # store contact in the notes field as structured text
        note = f"contact={contact_name}|linkedin={contact_linkedin}"
        app.notes = note
        s.commit(); s.close()
        self.log.info(f"Recorded application for job {job_id} ({status}).")
        return {"ok": True, "job_id": job_id}

    def rows(self) -> list[dict]:
        """All application records joined with job details, sheet-ready."""
        s = database.get_session()
        joined = (s.query(database.Job, database.Application)
                  .join(database.Application,
                        database.Job.id == database.Application.job_id).all())
        s.close()
        out = []
        for job, app in joined:
            contact, linkedin = "", ""
            if app.notes and "contact=" in app.notes:
                for part in app.notes.split("|"):
                    if part.startswith("contact="):
                        contact = part[8:]
                    elif part.startswith("linkedin="):
                        linkedin = part[9:]
            out.append({
                "Role": job.title, "Company": job.company,
                "Location": job.location, "Contact": contact,
                "Contact LinkedIn": linkedin,
                "Resume Used": app.cv_used or "", "Status": app.status,
                "Applied On": (app.updated_at.strftime("%Y-%m-%d")
                               if app.updated_at else ""),
                "Job URL": job.url,
            })
        return out

    def export_excel(self) -> str:
        """Write all records to a real .xlsx and return the path."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        rows = self.rows()
        wb = Workbook(); ws = wb.active; ws.title = "Applications"
        headers = ["Role", "Company", "Location", "Contact",
                   "Contact LinkedIn", "Resume Used", "Status",
                   "Applied On", "Job URL"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3A5F")
        for r in rows:
            ws.append([r[h] for h in headers])
        # sensible column widths
        widths = [26, 22, 22, 18, 32, 34, 12, 12, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        out = config.REPORTS / "applications.xlsx"
        wb.save(str(out))
        self.log.info(f"Exported {len(rows)} applications -> {out.name}")
        return str(out)
